# =============================================================================
# core/excel_handler.py — Import/Export Excel estilizado con paleta EA
# Proyecto: Digital VSM & Operational Intelligence - EA Innovation & Solutions
# =============================================================================

import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from config.settings import (
    COLOR_RED_MAIN,
    COLOR_BG_SECONDARY,
    COLOR_BORDER,
    COLOR_GOLD,
    COLOR_WHITE,
    COMPANY_NAME,
    REQUIRED_EXCEL_COLUMNS,
    OPTIONAL_EXCEL_COLUMNS,
    VALIDATION_RANGES,
)

# -----------------------------------------------------------------------------
# CONSTANTES INTERNAS DE ESTILO EXCEL
# -----------------------------------------------------------------------------
# openpyxl usa ARGB (sin '#') → prefijo FF para opacidad completa
_RED_FILL      = "FFD32F2F"   # COLOR_RED_MAIN
_DARK_FILL_1   = "FF1A1C24"   # COLOR_BG_SECONDARY (filas impares)
_DARK_FILL_2   = "FF2D2D3A"   # COLOR_BORDER (filas pares)
_GOLD_BORDER   = "FFFFC107"   # COLOR_GOLD
_WHITE_FONT    = "FFFFFFFF"   # COLOR_WHITE
_GOLD_FONT     = "FFFFC107"   # COLOR_GOLD
_GREY_FONT     = "FFCCCCCC"   # COLOR_TEXT_SECONDARY

# Definición de columnas de la plantilla
_TEMPLATE_COLUMNS = [
    ("name",             "Proceso",             20),
    ("cycle_time",       "C/T (s)",             12),
    ("changeover_time",  "C/O (s)",             12),
    ("wip",              "WIP (u)",             12),
    ("uptime",           "Uptime (%)",          14),
    ("process_order",    "Orden",               10),
    ("va_ratio",         "VA Ratio (0-1)",      16),
    ("operators",        "Operadores",          14),
    ("batch_size",       "Tamaño Lote",         14),
]


# -----------------------------------------------------------------------------
# FUNCIÓN AUXILIAR: construir borde dorado sutil
# -----------------------------------------------------------------------------
def _gold_border(style: str = "thin") -> Border:
    """Retorna un objeto Border con color dorado en los cuatro lados."""
    side = Side(style=style, color=_GOLD_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


# -----------------------------------------------------------------------------
# GENERACIÓN DE PLANTILLA EXCEL ESTILIZADA
# -----------------------------------------------------------------------------
@st.cache_data
def generate_excel_template() -> bytes:
    """
    Genera una plantilla Excel estilizada con la paleta EA para importar datos VSM.

    Estructura del archivo:
    - Hoja: 'VSM_Data'
    - Fila 1: Nombre empresa (merged cells, fondo rojo, texto blanco bold)
    - Fila 2: Headers de columnas (fondo rojo, texto blanco bold)
    - Filas 3+: Filas de ejemplo con celdas alternadas (#1A1C24 / #2D2D3A)
    - Bordes dorados sutiles en todo el rango de datos

    Returns:
        bytes: Contenido del archivo .xlsx listo para st.download_button
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "VSM_Data"

    num_cols = len(_TEMPLATE_COLUMNS)
    last_col_letter = get_column_letter(num_cols)

    # ------------------------------------------------------------------
    # FILA 1: Nombre de la empresa (merged, fondo rojo, texto blanco bold)
    # ------------------------------------------------------------------
    ws.merge_cells(f"A1:{last_col_letter}1")
    company_cell = ws["A1"]
    company_cell.value = f"  {COMPANY_NAME}  —  Digital VSM Data Template"
    company_cell.fill = PatternFill(
        fill_type="solid", fgColor=_RED_FILL
    )
    company_cell.font = Font(
        name="Calibri", bold=True, size=13, color=_WHITE_FONT
    )
    company_cell.alignment = Alignment(
        horizontal="center", vertical="center"
    )
    company_cell.border = _gold_border()
    ws.row_dimensions[1].height = 26

    # ------------------------------------------------------------------
    # FILA 2: Headers de columnas (fondo rojo, texto blanco bold)
    # ------------------------------------------------------------------
    header_fill = PatternFill(fill_type="solid", fgColor=_RED_FILL)
    header_font = Font(name="Calibri", bold=True, size=11, color=_WHITE_FONT)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (field, label, col_width) in enumerate(_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = _gold_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[2].height = 22

    # ------------------------------------------------------------------
    # FILAS DE EJEMPLO: datos de los 5 procesos fallback
    # ------------------------------------------------------------------
    example_rows = [
        ("Corte",     48,  120, 30, 95.0, 1, 0.70, 2, 1),
        ("Prensa",    52,   90, 25, 98.0, 2, 0.65, 1, 1),
        ("Ensamble",  52,  150, 40, 92.0, 3, 0.60, 3, 1),
        ("Prueba",    53,   60, 20, 99.0, 4, 0.80, 1, 1),
        ("Corrosión", 46,  180, 35, 96.0, 5, 0.55, 2, 1),
    ]

    for row_offset, row_data in enumerate(example_rows):
        excel_row = 3 + row_offset
        # Alternar fondo: impar → #1A1C24, par → #2D2D3A
        row_fill_color = _DARK_FILL_1 if row_offset % 2 == 0 else _DARK_FILL_2
        row_fill = PatternFill(fill_type="solid", fgColor=row_fill_color)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = _gold_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Columna A (nombre): texto blanco, alineado a la izquierda
            if col_idx == 1:
                cell.font = Font(name="Calibri", size=11, bold=True, color=_WHITE_FONT)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                cell.font = Font(name="Calibri", size=11, color=_GREY_FONT)

        ws.row_dimensions[excel_row].height = 20

    # ------------------------------------------------------------------
    # FILA DE INSTRUCCIONES debajo de los datos
    # ------------------------------------------------------------------
    instructions_row = 3 + len(example_rows) + 1
    ws.merge_cells(
        f"A{instructions_row}:{last_col_letter}{instructions_row}"
    )
    instr_cell = ws[f"A{instructions_row}"]
    instr_cell.value = (
        "Instrucciones: Agrega procesos a partir de la fila 3. "
        "Columnas requeridas: name, cycle_time, changeover_time, wip, uptime, process_order. "
        "No modifiques los encabezados. VA Ratio entre 0 y 1. Uptime entre 0 y 100."
    )
    instr_cell.fill = PatternFill(fill_type="solid", fgColor="FF0E1117")
    instr_cell.font = Font(name="Calibri", size=9, italic=True, color=_GOLD_FONT)
    instr_cell.alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True, indent=1
    )
    ws.row_dimensions[instructions_row].height = 30

    # ------------------------------------------------------------------
    # FREEZE panes: fijar las primeras 2 filas al hacer scroll
    # ------------------------------------------------------------------
    ws.freeze_panes = "A3"

    # ------------------------------------------------------------------
    # SERIALIZAR a bytes
    # ------------------------------------------------------------------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# -----------------------------------------------------------------------------
# IMPORTACIÓN Y VALIDACIÓN DE ARCHIVO EXCEL
# -----------------------------------------------------------------------------
def import_excel_file(uploaded_file) -> tuple[pd.DataFrame | None, list[str]]:
    """
    Lee y valida un archivo Excel subido por el usuario.

    Proceso de validación:
    1. Leer hoja 'VSM_Data' (o primera hoja si no existe)
    2. Verificar que existan las columnas requeridas
    3. Eliminar filas vacías (dropna en columnas requeridas)
    4. Convertir y coaccionar tipos de datos numéricos
    5. Validar rangos de valores
    6. Asignar columnas opcionales con defaults si no están presentes

    Args:
        uploaded_file: Objeto de archivo de st.file_uploader

    Returns:
        tuple: (DataFrame limpio | None, lista de mensajes de error/advertencia)
    """
    errors: list[str] = []
    warnings: list[str] = []

    # ------------------------------------------------------------------
    # 1. LEER ARCHIVO
    # ------------------------------------------------------------------
    try:
        # Intentar leer la hoja 'VSM_Data' primero
        try:
            df = pd.read_excel(uploaded_file, sheet_name="VSM_Data", engine="openpyxl")
        except Exception:
            # Fallback: leer la primera hoja disponible
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, sheet_name=0, engine="openpyxl")
            warnings.append(
                "Hoja 'VSM_Data' no encontrada — se usó la primera hoja del archivo."
            )
    except Exception as e:
        errors.append(f"No se pudo leer el archivo Excel: {e}")
        return None, errors

    # ------------------------------------------------------------------
    # 2. NORMALIZAR NOMBRES DE COLUMNAS
    # ------------------------------------------------------------------
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    # ------------------------------------------------------------------
    # 3. VERIFICAR COLUMNAS REQUERIDAS
    # ------------------------------------------------------------------
    missing_cols = [c for c in REQUIRED_EXCEL_COLUMNS if c not in df.columns]
    if missing_cols:
        errors.append(
            f"Columnas requeridas faltantes: {', '.join(missing_cols)}. "
            f"Asegúrate de usar la plantilla oficial."
        )
        return None, errors

    # ------------------------------------------------------------------
    # 4. ELIMINAR FILAS COMPLETAMENTE VACÍAS
    # ------------------------------------------------------------------
    df = df.dropna(subset=REQUIRED_EXCEL_COLUMNS, how="all")

    if df.empty:
        errors.append("El archivo no contiene datos de procesos (todas las filas están vacías).")
        return None, errors

    # ------------------------------------------------------------------
    # 5. ELIMINAR FILAS QUE TENGAN 'name' VACÍO
    # ------------------------------------------------------------------
    df = df[df["name"].notna() & (df["name"].astype(str).str.strip() != "")]

    if df.empty:
        errors.append("Ningún proceso tiene nombre válido en la columna 'name'.")
        return None, errors

    # ------------------------------------------------------------------
    # 6. CONVERSIÓN DE TIPOS NUMÉRICOS
    # ------------------------------------------------------------------
    numeric_cols = {
        "cycle_time":       int,
        "changeover_time":  int,
        "wip":              int,
        "uptime":           float,
        "process_order":    int,
        "va_ratio":         float,
        "operators":        int,
        "batch_size":       int,
    }

    for col, dtype in numeric_cols.items():
        if col not in df.columns:
            continue
        try:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            if dtype == int:
                df[col] = df[col].fillna(0).astype(int)
            else:
                df[col] = df[col].fillna(0.0).astype(float)
        except Exception as e:
            warnings.append(f"Advertencia en columna '{col}': {e}")

    # ------------------------------------------------------------------
    # 7. ELIMINAR FILAS CON VALORES NULOS EN COLUMNAS REQUERIDAS NUMÉRICAS
    # ------------------------------------------------------------------
    req_numeric = ["cycle_time", "changeover_time", "wip", "uptime", "process_order"]
    df = df.dropna(subset=[c for c in req_numeric if c in df.columns])

    if df.empty:
        errors.append(
            "No quedaron filas válidas después de limpiar valores nulos en columnas numéricas."
        )
        return None, errors

    # ------------------------------------------------------------------
    # 8. VALIDACIÓN DE RANGOS
    # ------------------------------------------------------------------
    invalid_rows: list[str] = []
    for col, (min_val, max_val) in VALIDATION_RANGES.items():
        if col not in df.columns:
            continue
        mask_out = (df[col] < min_val) | (df[col] > max_val)
        if mask_out.any():
            bad_rows = df.loc[mask_out, "name"].tolist()
            invalid_rows.append(
                f"'{col}' fuera de rango [{min_val}, {max_val}] en: {', '.join(str(r) for r in bad_rows)}"
            )

    if invalid_rows:
        for msg in invalid_rows:
            warnings.append(f"Advertencia de rango — {msg}")

    # ------------------------------------------------------------------
    # 9. ASIGNAR COLUMNAS OPCIONALES CON DEFAULTS
    # ------------------------------------------------------------------
    for col, default_val in OPTIONAL_EXCEL_COLUMNS.items():
        if col not in df.columns:
            df[col] = default_val
            warnings.append(
                f"Columna '{col}' no encontrada — se asignó valor por defecto: {default_val}"
            )

    # ------------------------------------------------------------------
    # 10. LIMPIAR NOMBRE DE PROCESO
    # ------------------------------------------------------------------
    df["name"] = df["name"].astype(str).str.strip()

    # ------------------------------------------------------------------
    # 11. ORDENAR POR process_order Y RESETEAR ÍNDICE
    # ------------------------------------------------------------------
    df = df.sort_values("process_order").reset_index(drop=True)

    # Consolidar todos los mensajes (primero warnings, sin errores fatales)
    all_messages = warnings  # errores ya habrían retornado antes

    return df, all_messages


# -----------------------------------------------------------------------------
# PREVIEW ESTILIZADO ANTES DE CONFIRMAR IMPORTACIÓN
# -----------------------------------------------------------------------------
def show_import_preview(df: pd.DataFrame) -> None:
    """
    Muestra un preview estilizado del DataFrame importado antes de que
    el usuario confirme la importación definitiva.

    Incluye:
    - Resumen de filas y columnas detectadas
    - st.dataframe con estilo oscuro y colores EA
    - Métricas rápidas: total procesos, C/T promedio, WIP total

    Args:
        df: DataFrame limpio retornado por import_excel_file()
    """
    if df is None or df.empty:
        st.warning("No hay datos para previsualizar.")
        return

    num_rows = len(df)
    num_cols = len(df.columns)

    # ------------------------------------------------------------------
    # RESUMEN RÁPIDO
    # ------------------------------------------------------------------
    st.markdown(
        f"""
        <div style="
            background-color: #1A1C24;
            border: 1px solid #FFC107;
            border-radius: 6px;
            padding: 10px 16px;
            margin-bottom: 12px;
        ">
            <span style="color:#FFC107; font-weight:bold; font-size:13px;">
                Vista previa de importación
            </span>
            <span style="color:#CCCCCC; font-size:12px; margin-left:16px;">
                {num_rows} proceso(s) detectado(s) &nbsp;·&nbsp; {num_cols} columna(s)
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------------------
    # MÉTRICAS RÁPIDAS
    # ------------------------------------------------------------------
    col1, col2, col3 = st.columns(3)

    avg_ct = df["cycle_time"].mean() if "cycle_time" in df.columns else 0
    total_wip = df["wip"].sum() if "wip" in df.columns else 0
    avg_uptime = df["uptime"].mean() if "uptime" in df.columns else 0

    with col1:
        st.metric("Total Procesos", num_rows)
    with col2:
        st.metric("C/T Promedio", f"{avg_ct:.1f} s")
    with col3:
        st.metric("WIP Total", int(total_wip))

    # ------------------------------------------------------------------
    # SELECCIONAR COLUMNAS RELEVANTES PARA EL PREVIEW
    # ------------------------------------------------------------------
    display_cols = [
        c for c in [
            "name", "cycle_time", "changeover_time",
            "wip", "uptime", "process_order", "va_ratio", "operators",
        ]
        if c in df.columns
    ]
    df_display = df[display_cols].copy()

    # Renombrar columnas para legibilidad en el preview
    col_rename = {
        "name":             "Proceso",
        "cycle_time":       "C/T (s)",
        "changeover_time":  "C/O (s)",
        "wip":              "WIP",
        "uptime":           "Uptime (%)",
        "process_order":    "Orden",
        "va_ratio":         "VA Ratio",
        "operators":        "Operadores",
    }
    df_display = df_display.rename(columns=col_rename)

    # ------------------------------------------------------------------
    # ESTILIZAR EL DATAFRAME CON PANDAS STYLER
    # ------------------------------------------------------------------
    def style_row(row):
        """Colorear filas alternadas con paleta EA."""
        idx = row.name
        bg = "#1A1C24" if idx % 2 == 0 else "#2D2D3A"
        return [f"background-color: {bg}; color: #CCCCCC;" for _ in row]

    styled_df = (
        df_display.style
        .apply(style_row, axis=1)
        .set_table_styles([
            {
                "selector": "thead th",
                "props": [
                    ("background-color", "#D32F2F"),
                    ("color", "#FFFFFF"),
                    ("font-weight", "bold"),
                    ("text-align", "center"),
                    ("border", "1px solid #FFC107"),
                ],
            },
            {
                "selector": "td",
                "props": [
                    ("border", "1px solid #2D2D3A"),
                    ("text-align", "center"),
                ],
            },
        ])
        .format({
            "C/T (s)":      "{:.0f}",
            "C/O (s)":      "{:.0f}",
            "WIP":          "{:.0f}",
            "Uptime (%)":   "{:.1f}",
            "VA Ratio":     "{:.2f}",
            "Operadores":   "{:.0f}",
        }, na_rep="—")
    )

    st.dataframe(styled_df, use_container_width=True, hide_index=True)

    # ------------------------------------------------------------------
    # ADVERTENCIA SOBRE COLUMNAS OPCIONALES FALTANTES
    # ------------------------------------------------------------------
    faltantes = [c for c in OPTIONAL_EXCEL_COLUMNS if c not in df.columns]
    if faltantes:
        st.caption(
            f"Columnas opcionales no encontradas y asignadas por defecto: "
            f"{', '.join(faltantes)}"
        )
